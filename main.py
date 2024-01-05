import bs4
import requests
import openpyxl


# Crear una url sin numero de pagina
url_base = 'http://books.toscrape.com/catalogue/page-{}.html'

# Lista de titulos con 4 o 5 estrellas
data = []

# Iterar paginas
for pagina in range(1, 51):

    # Crear sopa en cada pagina
    url_pagina = url_base.format(pagina)
    resultado = requests.get(url_pagina)
    sopa = bs4.BeautifulSoup(resultado.text, 'lxml')

    # Seleccionar datos de los libros
    libros = sopa.select('.product_pod')

    # Iterar los libros
    for libro in libros:

        # Revisar que tengan 4 o 5 estrellas
            if len(libro.select('.star-rating.Four')) != 0 or len(libro.select('.star-rating.Five')) != 0:

                # Guardar la información que necesitemos
                data.append((libro.select('a')[1]['title'], libro.select('.price_color')[0].getText()))


# Crear un archivo excel
workbook = openpyxl.Workbook()

# Seleccionar la hoja activa (por defecto, la primera hoja
sheet = workbook.active

# Escribir los encabezados en celdas específicas
sheet['A1'] = 'Nombre'
sheet['B1'] = 'Precio'

# Iterar sobre la lista y escribir en las celdas
for index, (nombre, precio) in enumerate(data, start=2):
    sheet.cell(row=index, column=1, value=nombre)
    sheet.cell(row=index, column=2, value=precio)

workbook.save('libros.xlsx')

